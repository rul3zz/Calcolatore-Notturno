
import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
import PyPDF2
import os
import subprocess
from datetime import datetime, time, timedelta

st.set_page_config(page_title="Gestione Turni V5.2", layout="wide")
st.title("🗓️ Gestione Turni V5.2 - Web Edition")
st.markdown("Carica il PDF dei turni e avvia l'elaborazione. Il file `meta.xlsx` sarà aggiornato automaticamente.")


def puliza_celle_iniziale():
    # Apri il file Excel esistente
    file_excel = "meta.xlsx"
    workbook = openpyxl.load_workbook(filename='meta.xlsx')
    sheet = workbook.active

    # Cancelliamo il contenuto delle celle dalla B5 alla B500
    for row in sheet.iter_rows(min_row=5, max_row=500, min_col=2, max_col=2):
        for cell in row:
            cell.value = None
    
    # Cancelliamo il contenuto delle celle dalla A5 alla A500
    for row in sheet.iter_rows(min_row=5, max_row=500, min_col=1, max_col=1):
        for cell in row:
            cell.value = None
    
    # Salva il file Excel modificato
    workbook.save(filename='meta.xlsx')


puliza_celle_iniziale()


def converti_orario(valore):
    """Converte un valore numerico in ore e minuti, assicurandosi che le ore rimangano entro il ciclo di 24 ore."""
    if isinstance(valore, (int, float)):
        ore = int(valore * 24) % 24  # Usa il modulo 24 per evitare valori oltre le 24:00
        minuti = round((valore * 1440) % 60)
        if minuti == 60:
            ore = (ore + 1) % 24
            minuti = 0
        return f"{ore:02}:{minuti:02}"
    return valore





####


def calcola_valore():
    # Apri il file Excel
    file_excel = "meta.xlsx"
    app = xw.App(visible=False)  # Excel in background
    wb = xw.Book(file_excel)  # Apri il file
    sheet = wb.sheets.active  # Seleziona il foglio attivo

    # Leggi il valore grezzo della cella H5
    valore_h5 = sheet.range("H5").value

    # Se il valore è numerico, convertilo in ore e minuti con arrotondamento
    if isinstance(valore_h5, (int, float)):
        ore = int(valore_h5 * 24)  # Parte intera -> ore
        minuti = round((valore_h5 * 1440) % 60)  # Arrotondamento corretto per evitare 59:59

        # Se i minuti arrivano a 60, incrementa le ore e resetta i minuti a 0
        if minuti == 60:
            ore += 1
            minuti = 0

        valore_h5 = f"{ore:02}:{minuti:02}"  # Formatta HH:MM

    # Chiudi il file Excel
    wb.close()
    app.quit()

    # Imposta il valore calcolato nella casella di testo (Entry)
    result_entry.config(state="normal")  # Attiva la modifica temporaneamente
    result_entry.delete(0, tk.END)  # Pulisce la casella
    result_entry.insert(0, valore_h5)  # Inserisce il valore calcolato
    result_entry.config(state="disabled")  # Disabilita la modifica















def pulizia_file():
  if os.path.exists("convertito_temp.xlsx"):
      os.remove("convertito_temp.xlsx")
  if os.path.exists("var.tmp"):
      os.remove("var.tmp")
  if os.path.exists("lib.tmp"):
      os.remove("lib.tmp")
  if os.path.exists("data.dat"):
      os.remove("data.dat")

# Alla pressione del tasto ESCI se presenti file cancellali e esci dal programma

def save_dates():
    start_date = start_date_entry.get_date()
    end_date = end_date_entry.get_date()
    if start_date > end_date:
        messagebox.showerror("Errore", "La data di inizio non può essere successiva alla data di fine.")
        exit ()
        tk_root.quit()
        return
    # Calcola la lista delle date comprese tra la data di inizio e la data di fine
    dates = []
    current_date = start_date
    while current_date <= end_date:
        dates.append(current_date.strftime("%d/%m/%Y"))
        current_date += timedelta(days=1)
    # Salva le date in un file di testo
    with open("lib.tmp", "w") as file:
        for date in dates:
            file.write(date + "\n")
    os.system('attrib +h "lib.tmp"')


#def chiedi_nome():
#    cognome_nome = ""
#    while not cognome_nome:  # Rimane bloccato finché l'input è vuoto
#        cognome_nome = input("Inserisci il cognome e nome del dipendente: ").strip()
    
#    return cognome_nome  # Restituisce il valore inserito quando è valido




def avvia_elaborazione():
    """Esegue le funzioni solo se il campo non è vuoto."""
    if nome_entry.get().strip():  # Se il campo non è vuoto
        save_dates()
        process_file()
        process_profilo_orario()
        finale()
        calcola_valore()
        abilita_mostra()
    else:
        messagebox.showwarning("Attenzione", "Inserisci il cognome e nome prima di procedere.")  # Opzionale


def process_file():
    
    # Esegui la funzione per killare il processo "meta.xlsx"
    kill_process("EXCEL.EXE")
    
    # ottiene il percorso della cartella di lavoro
    cartella_di_lavoro = os.getcwd()
    cognome_nome = nome_entry.get().strip().upper()
    
    # Lettura del PDF caricato
    with open(pdf_path, 'rb') as pdf_file:
        pdf_reader = PyPDF2.PdfReader(pdf_file)
        text = "".join([pdf_reader.pages[page].extract_text() for page in range(len(pdf_reader.pages))])
    # elabora il testo in un DataFrame di Pandas
    rows = []
    for line in text.split('\n'):
        cells = line.split(' ')
        if len(cells) > 2:
            if len(cells[2]) > 4:
                cells[0] = ' '.join(cells[:3])    ## Occhio a questa parte è stata modificata in maniera strana nel passaggio a tkinter
                del cells[1:3]    ## Occhio a questa parte è stata modificata in maniera strana nel passaggio a tkinter
            else:
                cells[0] = ' '.join(cells[:2])    ## Occhio a questa parte è stata modificata in maniera strana nel passaggio a tkinter
                del cells[1]    ## Occhio a questa parte è stata modificata in maniera strana nel passaggio a tkinter
        rows.append(cells)
    df = pd.DataFrame(rows)
    # esporta il DataFrame in un file Excel
    df.to_excel("convertito_temp.xlsx", index=False)
    # nome del file Excel da leggere
    xlsx_files = [f for f in glob.glob(cartella_di_lavoro + "/*.xlsx") if os.path.isfile(f) and os.path.basename(f) != "meta.xlsx"]
    if not xlsx_files:
        messagebox.showerror("Errore #122", "File di sistema mancanti.")   ## Vecchio messaggio Nessun file XLSX trovato nella cartella.
        tk_root.quit()
        return
    nome_file = xlsx_files[0]
    workbook = openpyxl.load_workbook(nome_file)
    sheet = workbook.active
    os.system('attrib +h "convertito_temp.xlsx"')
#    messagebox.showinfo("Successo", "Elaborazione completata con successo!")





    # Cancella il contenuto dei file
    file_names = ["data.dat", "var.tmp"]
    for file_name in file_names:
        try:
            with open(file_name, "w") as f:
                f.truncate(0)
        except:
            pass
    

    xlsx_files = [f for f in glob.glob(cartella_di_lavoro + "/*.xlsx") if os.path.isfile(f) and os.path.basename(f) != "meta.xlsx"]
    if not xlsx_files:
        messagebox.showerror("Errore #182", "File di sistema mancanti.") #Nessun file XLSX trovato nella cartella.
        return
    
    nome_file = xlsx_files[0]
    workbook = openpyxl.load_workbook(nome_file)
    sheet = workbook.active
    # cerca il nome e il cognome nella colonna A
    trovato = False
    for row in sheet.iter_rows(min_row=2):
        nome_cognome = row[0].value
        if nome_cognome == cognome_nome:
            valori = [cell.value for cell in row[1:]]
            trovato = True
            break
    
    # crea un file di testo con i valori trovati in colonna, ignorando i valori vuoti
    if trovato:
        with open("data.dat", "w") as f:
            for v in valori:
                if v is not None:
                    if v.startswith("R"):
                        if len(v) > 1 and (v[1] == '.' or v[1] == '$'):
                            f.write("{}{}\n".format(v[0], v[1]))
                            f.write("{}\n".format(v[2:]))
                        else:
                            f.write("R\n")
                            f.write("{}\n".format(v[1:]))
                    else:
                        f.write("{}\n".format(v))
        # Rimuovi righe vuote nel file
        with open("data.dat", "r+") as f:
            lines = f.readlines()
            f.seek(0)
            f.writelines(line for line in lines if line.strip())
            f.truncate()
        
        messagebox.showinfo("Successo", f"Dipendente {cognome_nome} trovato. Calcolo Eseguito")
    else:
        messagebox.showerror("Errore", "Nessun dipendente trovato per il cognome e nome inseriti.")
        os.remove("convertito_temp.xlsx")
        os.remove("lib.tmp")
        os.remove("var.tmp")
        os.remove("data.dat")
        tk_root.quit()
        return





def process_profilo_orario():
    profilo = profilo_var.get()
    file_nascosto_convertiti_per_calcolo_notturno = "var.tmp"
    workbook = openpyxl.load_workbook(filename='meta.xlsx')
    sheet = workbook.active
    os.system(f'attrib -h "{file_nascosto_convertiti_per_calcolo_notturno}"')
    
    if profilo == "Part Time 5 Ore":
        orari = {"4T": "04:00", "43T": "04:30", "5T": "05:00", "53T": "05:30", "63T": "06:30", "7T": "07:00", "73T": "07:30", "8T": "08:00", "83T": "08:30", "9T": "09:00", "93T": "09:30", "10T": "10:00", "103T": "10:30", "11T": "11:00", "113T": "11:30", "12T": "12:00", "123T": "12:30", "13T": "13:00", "133T": "13:30", "14T": "14:00", "143T": "14:30", "15T": "15:00", "153T": "15:30", "16T": "16:00", "163T": "16:30", "17T": "17:00", "173T": "17:30", "18T": "18:00", "183T": "18:30", "19T": "19:00", "24T": "00:00"}
        esclusi = {"GF", "R", "F", "R.", "R$", "CIG", "CS", "XX", "CDG", "CPG", "CPS", "FTA", "PST", "ROL", "PNR", "PS", "PLT", "PEL", "PCE", "AG", "AGNR", "CDO", "D.S.", "D.P.", "INF", "L104", "MA", "MFG", "MTA", "FMT", "SOS", "SOSC"}
        with open("data.dat", "r") as fp_in, open("var.tmp", "w") as fp_out:
            for line in map(str.strip, fp_in):
                if line.endswith("T") and line not in esclusi:
                    fp_out.write(f"{orari.get(line, '')}\n")
        sheet['F2'] = "5:00"
    
    elif profilo == "Part Time 6 Ore":
        time_mapping = {"4": "04:00", "43": "04:30", "5": "05:00", "53": "05:30", "63": "06:30", "7": "07:00", "73": "07:30", "8": "08:00", "83": "08:30", "9": "09:00", "93": "09:30", "10": "10:00", "103": "10:30", "11": "11:00", "113": "11:30", "12": "12:00", "123": "12:30", "13": "13:00", "133": "13:30", "14": "14:00", "143": "14:30", "15": "15:00", "153": "15:30", "16": "16:00", "163": "16:30", "17": "17:00", "173": "17:30", "18": "18:00", "183": "18:30", "19": "19:00", "23": "23:00", "24": "00:00"}
        invalid_values = {"GF", "R", "F", "R.", "R$", "CIG", "CS", "XX", "CDG", "CPG", "CPS", "FTA", "PST", "ROL", "PNR", "PS", "PLT", "PEL", "PCE", "AG", "AGNR", "CDO", "D.S.", "D.P.", "INF", "L104", "MA", "MFG", "MTA", "FMT", "SOS", "SOSC"}
        with open("data.dat", "r") as fp_in, open("var.tmp", "w") as fp_out:
            for line in fp_in:
                line = line.strip()
                if line not in invalid_values and line in time_mapping:
                    fp_out.write(time_mapping[line] + "\n")
        sheet['F2'] = "6:00"
    
    elif profilo == "Full Time":
        orari = {"K": "04:30", "W": "05:00", "C": "05:30", "A": "06:00", "V": "07:00", "X": "08:00", "T": "09:00", "Y": "10:00", "S": "11:00", "Z": "11:48", "Z3": "12:18", "U": "12:48", "J": "13:18", "B": "13:48", "B3": "14:18", "D": "14:48", "D3": "15:18", "E": "15:48", "E3": "16:18", "H": "16:48", "I": "17:48", "L3": "21:30", "N": "22:00"}
        esclusi = {"GF", "R", "F", "R.", "R$", "CIG", "CS", "XX", "CDG", "CPG", "CPS", "FTA", "PST", "ROL", "PNR", "PS", "PLT", "PEL", "PCE", "AG", "AGNR", "CDO", "D.S.", "D.P.", "INF", "L104", "MA", "MFG", "MTA", "FMT", "SOS", "SOSC"}
        with open("data.dat", "r") as fp_in, open("var.tmp", "w") as fp_out:
            for line in fp_in:
                line = line.strip()
                if line not in esclusi and line in orari:
                    fp_out.write(orari[line] + "\n")
        sheet['F2'] = "8:12"
    
    workbook.save(filename='meta.xlsx')






def finale():
    # Apri il file Excel esistente
    workbook = openpyxl.load_workbook(filename='meta.xlsx')
    sheet = workbook.active

    # Cancelliamo il contenuto delle celle dalla B5 alla B500
    for row in sheet.iter_rows(min_row=5, max_row=500, min_col=2, max_col=2):
        for cell in row:
            cell.value = None
    
    # Cancelliamo il contenuto delle celle dalla A5 alla A500
    for row in sheet.iter_rows(min_row=5, max_row=500, min_col=1, max_col=1):
        for cell in row:
            cell.value = None
    
    # Leggi i valori dal file "var.tmp"
    with open('var.tmp', 'r') as f:
        values = [line.strip() for line in f]

    # Modifica i valori per ottenere un formato compatibile con Excel (hh:mm -> hh:mm:ss)
    values = [value + ':00' for value in values]

    # Inserisci i valori nella colonna B del file Excel a partire dalla cella B5 come oggetti time
    for i, value in enumerate(values):
        t = datetime.strptime(value, '%H:%M:%S').time()
        cell = sheet.cell(row=i+5, column=2)
        cell.value = t
    
    # Imposta il formato della colonna B come ora
    sheet.column_dimensions['B'].number_format = 'hh:mm'

    # Leggi le date dal file "lib.tmp"
    with open('lib.tmp', 'r') as date_file:
        dates = date_file.readlines()

    # Leggi i valori dai file "data.dat"
    with open('data.dat', 'r') as values_file:
        values = values_file.readlines()

    # Riga iniziale del foglio Excel
    row = 5

    # Itera su ogni data e valore
    for date, value in zip(dates, values):
        date = date.strip()  # Rimuovi eventuali spazi o caratteri di fine riga
        value = value.strip()
        
        # Se il valore non è nella lista di quelli da ignorare
        if value not in ["GF", "R", "F", "R.", "R$", "CIG", "CS", "XX", "CDG", "CPG", "CPS", "FTA"]:
            # Inserisci la data nella Colonna A
            sheet[f'A{row}'] = date
            row += 1

    # Salva il file Excel modificato
    workbook.save(filename='meta.xlsx')

    # Rimuove i file temporanei
    os.remove("convertito_temp.xlsx")
    os.remove("lib.tmp")
    os.remove("var.tmp")
    
    # Pulisce la console e mostra un messaggio di completamento
    os.system('cls' if os.name == 'nt' else 'clear')




# Caricamento PDF
pdf_file = st.file_uploader("📄 Carica il file PDF dei turni", type=["pdf"])

if st.button("🚀 Avvia Elaborazione"):
    if not pdf_file:
        st.warning("Per favore carica un file PDF.")
    else:
        try:
            # Salvataggio temporaneo del PDF
            with open("turni_input.pdf", "wb") as f:
                f.write(pdf_file.read())

            # Esegui pipeline logica
            pulizia_file()
            puliza_celle_iniziale()
            process_file()  # parsing PDF e preparazione dati
            process_profilo_orario()
            calcola_valore()
            save_dates()
            finale()  # salva su meta.xlsx

            st.success("✅ Elaborazione completata! Il file `meta.xlsx` è stato aggiornato.")
            with open("meta.xlsx", "rb") as f:
                st.download_button("📥 Scarica meta.xlsx aggiornato", data=f, file_name="meta_aggiornato.xlsx")

        except Exception as e:
            st.error(f"Errore durante l'elaborazione: {e}")
